import io
from decimal import Decimal
from datetime import datetime
from openpyxl import Workbook

def clean_number(raw_val):
    """
    - None -> ""
    - Decimal/float/int -> string tanpa trailing .000
    - string -> trim
    """
    if raw_val is None:
        return ""
    if isinstance(raw_val, (int, float, Decimal)):
        s = format(Decimal(str(raw_val)), 'f')
    else:
        s = str(raw_val).strip()
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s

def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return v or ""

def merge_stok_into_master(master_rows, stok_pre_rows, stok_post_rows):
    """
    Gabungkan:
    - master_rows         -> info barang, ruangan, harga
    - stok_pre_rows       -> stok sistem sebelum SO (STOK_PRE_SO)
    - stok_post_rows (SO final)  -> stok hasil opname fisik (STOK_SO_FINAL)
                           -> stok setelah SO (STOK_POST_SO) -> diasumsikan sama

    Hasil akhir per barang akan punya kolom:
    STOK_PRE_SO, STOK_SO_FINAL, STOK_POST_SO,
    plus selisih dan timestamp.
    """

    # map stok sebelum SO, key = ID_BARANG
    pre_map = { row["ID_BARANG"]: row for row in stok_pre_rows }

    # map stok final SO, key = ID_BARANG
    so_map = { row["ID_BARANG"]: row for row in stok_post_rows }

    merged = []
    for m in master_rows:
        barang_id = m["ID_BARANG"]

        # --- stok awal sistem sebelum opname (cutoff 23:59:59)
        if barang_id in pre_map:
            stok_pre_so = pre_map[barang_id]["STOK"]
            tanggal_cutoff = pre_map[barang_id]["TANGGAL"]
        else:
            stok_pre_so = m.get("STOK", 0)
            tanggal_cutoff = m.get("TANGGAL")

        # --- stok setelah SO = hasil fisik opname
        if barang_id in so_map:
            stok_post_so = so_map[barang_id]["STOK_POST_SO"]
            tanggal_post_so = so_map[barang_id]["TANGGAL_POST_SO"]
        else:
            stok_post_so = 0
            tanggal_post_so = '0000-00-00'

        # --- hitung selisih
        try:
            diff_pre_vs_post = Decimal(str(stok_post_so)) - Decimal(str(stok_pre_so))
        except Exception:
            diff_pre_vs_post = ""

        # row final
        row = {
            "ID_BARANG": m["ID_BARANG"],
            "NAMA": m["NAMA"],
            "SATUAN": m["SATUAN"],
            "KATEGORI": m["KATEGORI"],
            "HARGA_JUAL": m.get("HARGA_JUAL"),

            "STOK_PRE_SO": stok_pre_so,
            "STOK_POST_SO": stok_post_so,

            "SELISIH_PRE_VS_POST": diff_pre_vs_post,

            "ID_BARANG_RUANGAN": m["ID_BARANG_RUANGAN"],
            "KODE_RUANGAN": m["KODE_RUANGAN"],
            "RUANGAN": m["RUANGAN"],
            "STATUS_BARANG": m["STATUS_BARANG"],
            "STATUS_BARANG_RUANGAN": m["STATUS_BARANG_RUANGAN"],

            "TANGGAL_CUTOFF": tanggal_cutoff,
            "TANGGAL_POST_SO": tanggal_post_so,
        }

        merged.append(row)

    return merged

def build_xlsx_so(merged_rows, stok_pre_rows, stok_post_rows):
    wb = Workbook()

    # MASTER_DATA
    ws1 = wb.active
    ws1.title = "MASTER_DATA"
    headers1 = [
        "ID_BARANG",
        "NAMA",
        "SATUAN",
        "KATEGORI",
        "HARGA_JUAL",

        "STOK_PRE_SO",
        "STOK_POST_SO",

        "SELISIH_PRE_VS_POST",

        "ID_BARANG_RUANGAN",
        "KODE_RUANGAN",
        "RUANGAN",
        "STATUS_BARANG",
        "STATUS_BARANG_RUANGAN",

        "TANGGAL_CUTOFF",
        "TANGGAL_POST_SO",
    ]
    ws1.append(headers1)

    for r in merged_rows:
        ws1.append([
            r["ID_BARANG"],
            r["NAMA"],
            r["SATUAN"],
            r["KATEGORI"],
            clean_number(r["HARGA_JUAL"]),

            clean_number(r["STOK_PRE_SO"]),
            clean_number(r["STOK_POST_SO"]),

            clean_number(r["SELISIH_PRE_VS_POST"]),

            r["ID_BARANG_RUANGAN"],
            r["KODE_RUANGAN"],
            r["RUANGAN"],
            r["STATUS_BARANG"],
            r["STATUS_BARANG_RUANGAN"],

            fmt_date(r["TANGGAL_CUTOFF"]),
            fmt_date(r["TANGGAL_POST_SO"]),
        ])

    # STOK_PRA_SO
    ws2 = wb.create_sheet("STOK_PRE_SO")
    headers2 = [
        "ID_BARANG_RUANGAN",
        "KODE_RUANGAN",
        "RUANGAN",
        "ID_BARANG",
        "NAMA",
        "SATUAN",
        "KATEGORI",
        "HARGA_JUAL",
        "STOK",
        "STATUS_BARANG_RUANGAN",
        "STATUS_BARANG",
        "TANGGAL_CUTOFF",
    ]
    ws2.append(headers2)

    for s in stok_pre_rows:
        ws2.append([
            s["ID_BARANG_RUANGAN"],
            s["KODE_RUANGAN"],
            s["RUANGAN"],
            s["ID_BARANG"],
            s["NAMA"],
            s["SATUAN"],
            s["KATEGORI"],
            s["HARGA_JUAL"],
            clean_number(s["STOK"]),
            s["STATUS_BARANG_RUANGAN"],
            s["STATUS_BARANG"],
            fmt_date(s["TANGGAL"]),
        ])

    # STOK_POST_SO
    ws3 = wb.create_sheet("STOK_POST_SO")
    headers3 = [
        "IDSO",
        "STOK_OPNAME",
        "BARANG_RUANGAN",
        "ID_BARANG",
        "NAMA_BARANG",
        "MANUAL_STOK_POST_SO",
        "HARGA_BELI+PPN",
        "KODE_RUANGAN",
        "RUANGAN",
        "TANGGAL_POST_SO",
    ]
    ws3.append(headers3)

    for z in stok_post_rows:
        ws3.append([
            z["IDSO"],
            z["STOK_OPNAME"],
            z["BARANG_RUANGAN"],
            z["ID_BARANG"],
            z["NAMA_BARANG"],
            clean_number(z["STOK_POST_SO"]),
            z["HARGA_JUAL"],            
            z["KODE_RUANGAN"],
            z["RUANGAN"],
            fmt_date(z["TANGGAL_POST_SO"]),
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def build_ed_xlsx(rows, threshold_days):
    wb = Workbook()
    ws = wb.active
    ws.title = "OBAT_EXPIRED_DATE"

    headers = [
        "ID_BARANG",
        "NAMA",
        "NO_FAKTUR",
        "NO_BATCH",
        "TGL_EXP",
        "STATUS_EXP",
        "STOK_DITERIMA",        
        "KATEGORI",
        "SATUAN",
        "HARGA_SATUAN",
        "TOTAL_HARGA",
        "REKANAN",
        "TGL_PENERIMAAN",
        "STATUS_BARANG"
    ]
    ws.append(headers)

    def to_datestr(val):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        return "" if val is None else str(val)

    for r in rows:
        ws.append([
            r["ID"],
            r["NAMABARANG"],       
            r["NOFAKTUR"],        
            r["NO_BATCH"],            
            to_datestr(r["TGL_EXP"]),
            r["STATUS_EXP"],
            r["STOK_DITERIMA"],
            r["KATEGORI"],
            r["NAMASATUAN"],
            r["HARGASATUAN"],
            r["TOTAL_HARGA"],
            r["NAMAREKANAN"],
            to_datestr(r["TANGGAL_PENERIMAAN"]),
            r["STATUS"],
        ])

    # Info summary kecil di sheet kedua
    ws2 = wb.create_sheet("INFO")
    ws2["A1"] = "Laporan EXP"
    ws2["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws2["A2"] = "Hari EXP (hari) setelah " + datetime.now().strftime("%Y-%m-%d")
    ws2["B2"] = threshold_days

    ws2["A3"] = "Total item"
    ws2["B3"] = len(rows)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio